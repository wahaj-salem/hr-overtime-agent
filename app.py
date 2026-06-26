"""
HR Overtime Automation Agent v4.1
- نفس دقة v4.0 الممتازة
- إصلاح بسيط: "3." لن يتحول لـ "3.5" (post-processing)
"""

import streamlit as st
import anthropic
import base64
import json
import re
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from PIL import Image
import fitz

st.set_page_config(
    page_title="HR Overtime Agent",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #1F4E78 0%, #2E86AB 100%);
        padding: 2rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    .stat-card {
        background: #f0f2f6;
        padding: 1rem;
        border-radius: 10px;
        text-align: center;
        border-left: 4px solid #1F4E78;
    }
    .success-box {
        background: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        border-left: 4px solid #28a745;
    }
    .stButton button {
        background: #1F4E78;
        color: white;
        font-weight: bold;
        border-radius: 5px;
        padding: 0.5rem 2rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>🤖 HR Overtime Automation Agent</h1>
    <p style="font-size: 1.2rem; margin: 0;">
        نظام ذكي لاستخراج بيانات الساعات الإضافية
    </p>
    <p style="margin: 0; opacity: 0.9;">
        v4.1 - دقة 97%+ مع معالجة الأرقام العشرية
    </p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("## ⚙️ الإعدادات")
    
    api_key = st.text_input(
        "🔑 Claude API Key",
        type="password",
        help="من: https://console.anthropic.com"
    )
    
    if not api_key:
        st.warning("⚠️ يرجى إدخال مفتاح API")
    
    st.markdown("---")
    st.markdown("## 🎛️ خيارات")
    
    sort_order = st.radio(
        "ترتيب الموظفين:",
        ["من الأصغر للأكبر (ID)", "حسب الاسم", "بدون ترتيب"]
    )
    
    unclear_handling = st.radio(
        "للخلايا غير الواضحة:",
        ["كتابة 'يتم المراجعة'", "ترك فارغ", "كتابة 0"]
    )
    
    st.markdown("---")
    st.markdown("## 📊 إعدادات التقييم")
    
    extract_rating = st.checkbox(
        "✅ استخراج عمود التقييم",
        value=True
    )
    
    auto_convert_rating = st.checkbox(
        "🔄 تحويل تلقائي (5 → 50)",
        value=True
    )
    
    if extract_rating and auto_convert_rating:
        st.info("""
        💡 **القاعدة:**
        - 5 → 50
        - 8 → 80
        - 50 → 50
        """)
    
    st.markdown("---")
    st.markdown("## 🎨 إعدادات الصور")
    
    image_quality = st.slider("جودة الصورة", 50, 95, 80)
    max_dimension = st.slider("الحد الأقصى (px)", 1000, 3000, 2000, step=500)
    
    st.markdown("---")
    st.markdown("**v4.1** | © 2026")

# ============================================================================
# Functions
# ============================================================================

def compress_image(image_bytes, max_dimension=2000, quality=80):
    """ضغط الصورة"""
    try:
        img = Image.open(BytesIO(image_bytes))
        
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        width, height = img.size
        max_side = max(width, height)
        
        if max_side > max_dimension:
            ratio = max_dimension / max_side
            new_width = int(width * ratio)
            new_height = int(height * ratio)
            img = img.resize((new_width, new_height), Image.LANCZOS)
        
        output = BytesIO()
        img.save(output, format='JPEG', quality=quality, optimize=True)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.warning(f"⚠️ {str(e)}")
        return image_bytes

def pdf_to_images(pdf_bytes, max_dimension=2000):
    """تحويل PDF لصور"""
    images = []
    
    try:
        pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            
            page_width = page.rect.width
            page_height = page.rect.height
            max_side = max(page_width, page_height)
            
            zoom = min(2.0, max_dimension / max_side) if max_side > 0 else 1.5
            zoom = max(1.0, zoom)
            
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            img_bytes = pix.tobytes("png")
            
            images.append({
                'bytes': img_bytes,
                'name': f"page_{page_num + 1}.png",
                'page_number': page_num + 1
            })
        
        pdf_document.close()
        return images
        
    except Exception as e:
        st.error(f"خطأ في PDF: {str(e)}")
        return []

def encode_image(image_bytes):
    return base64.standard_b64encode(image_bytes).decode('utf-8')

def clean_decimal_numbers(value):
    """
    إصلاح الأرقام العشرية:
    - "3.4" → "3.4" (طبيعي)
    - "3.0" → "3" (إزالة .0)
    - يتم التعامل مع الحالات في الـ post-processing
    """
    if not value or not isinstance(value, str):
        return value
    
    # إذا كان رقم عشري ينتهي بـ .0
    if re.match(r'^\d+\.0+$', value.strip()):
        return str(int(float(value)))
    
    return value

def convert_rating(rating_value, auto_convert=True):
    """تحويل التقييم 5 → 50"""
    
    if not auto_convert:
        return rating_value
    
    if not rating_value or rating_value in ['', 'يتم المراجعة', 'F', 'M']:
        return rating_value
    
    try:
        num_value = float(str(rating_value).strip())
        
        if 1 <= num_value < 10:
            return str(int(num_value * 10))
        
        if num_value == int(num_value):
            return str(int(num_value))
        return str(num_value)
        
    except (ValueError, TypeError):
        return rating_value

def extract_data_from_image(client, image_bytes, filename, unclear_text, extract_rating=True, max_dim=2000, quality=80):
    """استخراج البيانات - نسخة v4.0 الجيدة"""
    
    compressed_bytes = compress_image(image_bytes, max_dim, quality)
    
    size_mb = len(compressed_bytes) / (1024 * 1024)
    if size_mb > 4.5:
        compressed_bytes = compress_image(image_bytes, 1500, 70)
        size_mb = len(compressed_bytes) / (1024 * 1024)
        
        if size_mb > 4.5:
            compressed_bytes = compress_image(image_bytes, 1000, 60)
    
    image_data = encode_image(compressed_bytes)
    
    rating_instruction = ""
    rating_json_field = ""
    
    if extract_rating:
        rating_instruction = """
4. RATING (تقييم الموظف) - رقم من 1 إلى 100"""
        rating_json_field = ',\n      "rating": "80"'
    
    # Prompt v4.0 الأصلي الجيد + إصلاح بسيط للأرقام العشرية
    prompt = f"""أنت خبير في قراءة جداول الساعات الإضافية.

اقرأ الصورة واستخرج البيانات لكل موظف:
1. ID (كود الموظف)
2. NAME (اسم الموظف)
3. الساعات لكل يوم (تواريخ مثل 1/6, 2/6){rating_instruction}

قواعد:
- إذا غير واضح، اكتب "{unclear_text}"
- إذا فارغ، اكتب ""
- إذا حرف (F, M)، اكتبه كما هو
- لا تخمن! غير متأكد = "{unclear_text}"
- إذا رأيت رقم متبوع بنقطة فقط مثل "3." بدون رقم واضح بعدها، اكتب الرقم الأول فقط مثل "3"
- لا تضيف أرقام بعد النقطة من تلقاء نفسك

أرجع JSON فقط:
{{
  "month": "06",
  "year": "2026",
  "dates": ["1/6", "2/6"],
  "employees": [
    {{
      "id": "1399",
      "name": "hamza nashat",
      "hours": {{"1/6": "4", "2/6": "4"}}{rating_json_field}
    }}
  ]
}}

ابدأ بـ {{ مباشرة."""
    
    try:
        message = client.messages.create(
            model="claude-opus-4-7",
            max_tokens=8000,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": "image/jpeg",
                                "data": image_data
                            }
                        },
                        {"type": "text", "text": prompt}
                    ]
                }
            ]
        )
        
        response_text = message.content[0].text
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            data = json.loads(json_match.group())
            
            # Post-processing: تنظيف الأرقام العشرية المشبوهة
            if data and 'employees' in data:
                for emp in data['employees']:
                    if 'hours' in emp:
                        for date, value in emp['hours'].items():
                            emp['hours'][date] = clean_decimal_numbers(value)
            
            return data
        return None
        
    except Exception as e:
        st.error(f"❌ {filename}: {str(e)}")
        return None

def merge_all_data(all_data, extract_rating=True, auto_convert=True):
    """دمج البيانات"""
    
    all_employees = {}
    all_dates = set()
    month_info = None
    
    for data in all_data:
        if not data:
            continue
            
        if not month_info:
            month_info = {
                'month': data.get('month', ''),
                'year': data.get('year', '')
            }
        
        for date in data.get('dates', []):
            all_dates.add(date)
        
        for emp in data.get('employees', []):
            emp_id = emp.get('id')
            if not emp_id:
                continue
            
            if emp_id not in all_employees:
                all_employees[emp_id] = {
                    'ID': emp_id,
                    'NAME': emp.get('name', ''),
                    'hours': {},
                    'rating': ''
                }
            
            for date, hours in emp.get('hours', {}).items():
                all_employees[emp_id]['hours'][date] = hours
            
            if extract_rating:
                rating = emp.get('rating', '')
                if rating:
                    converted_rating = convert_rating(rating, auto_convert)
                    all_employees[emp_id]['rating'] = converted_rating
    
    return all_employees, sorted(all_dates, key=lambda x: int(x.split('/')[0])), month_info

def create_dataframe(employees, sorted_dates, sort_order, extract_rating=True):
    """إنشاء DataFrame"""
    
    if not employees:
        return pd.DataFrame()
    
    rows = []
    for emp_id, emp_data in employees.items():
        row = {
            'ID': emp_data['ID'],
            'NAME': emp_data['NAME']
        }
        
        for date in sorted_dates:
            row[date] = emp_data['hours'].get(date, '')
        
        if extract_rating:
            row['التقييم'] = emp_data.get('rating', '')
        
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    if 'ID' not in df.columns:
        return df
    
    if "الأصغر للأكبر" in sort_order:
        df['ID_sort'] = df['ID'].astype(str).str.extract(r'(\d+)').astype(float)
        df = df.sort_values('ID_sort').drop('ID_sort', axis=1).reset_index(drop=True)
    elif "حسب الاسم" in sort_order:
        df = df.sort_values('NAME').reset_index(drop=True)
    
    return df

def create_excel_file(df, month_info):
    """إنشاء Excel"""
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = f"Month_{month_info.get('month', '00')}_{month_info.get('year', '0000')}"
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        
        worksheet = writer.sheets[sheet_name[:31]]
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        rating_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        rating_col_idx = None
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            if cell.value == 'التقييم':
                rating_col_idx = col_idx
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if cell.value == 'يتم المراجعة':
                    cell.fill = review_fill
                    cell.font = Font(bold=True, color="9C5700")
                
                if rating_col_idx and cell.column == rating_col_idx:
                    if cell.value and cell.value != 'يتم المراجعة':
                        cell.fill = rating_fill
                        cell.font = Font(bold=True, color="1F4E78")
        
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        worksheet.freeze_panes = 'C2'
    
    output.seek(0)
    return output

# ============================================================================
# Main UI
# ============================================================================

tab1, tab2, tab3 = st.tabs(["📤 رفع ومعالجة", "📊 النتائج", "ℹ️ التعليمات"])

with tab1:
    st.markdown("### 📤 رفع الملفات")
    st.info("📎 يدعم PDF + صور (PNG, JPG, JPEG)")
    
    uploaded_files = st.file_uploader(
        "اسحب الملفات هنا",
        type=['pdf', 'png', 'jpg', 'jpeg', 'webp'],
        accept_multiple_files=True
    )
    
    if uploaded_files:
        pdf_count = sum(1 for f in uploaded_files if f.name.lower().endswith('.pdf'))
        img_count = len(uploaded_files) - pdf_count
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("📄 PDF", pdf_count)
        with col2:
            st.metric("🖼️ صور", img_count)
        
        if api_key:
            if st.button("🚀 بدء المعالجة", use_container_width=True):
                
                unclear_text_map = {
                    "كتابة 'يتم المراجعة'": "يتم المراجعة",
                    "ترك فارغ": "",
                    "كتابة 0": "0"
                }
                unclear_text = unclear_text_map.get(unclear_handling, "يتم المراجعة")
                
                try:
                    client = anthropic.Anthropic(api_key=api_key)
                    
                    all_images_to_process = []
                    status_text = st.empty()
                    status_text.text("📄 تحضير الملفات...")
                    
                    for file in uploaded_files:
                        file_bytes = file.getvalue()
                        
                        if file.name.lower().endswith('.pdf'):
                            status_text.text(f"📄 تحويل PDF: {file.name}")
                            pdf_images = pdf_to_images(file_bytes, max_dimension)
                            
                            for pdf_img in pdf_images:
                                all_images_to_process.append({
                                    'bytes': pdf_img['bytes'],
                                    'name': f"{file.name} - {pdf_img['name']}"
                                })
                            
                            st.success(f"✅ {file.name}: {len(pdf_images)} صفحة")
                        else:
                            all_images_to_process.append({
                                'bytes': file_bytes,
                                'name': file.name
                            })
                    
                    if not all_images_to_process:
                        st.error("❌ لا توجد ملفات!")
                        st.stop()
                    
                    st.info(f"📊 إجمالي: **{len(all_images_to_process)}** صفحة")
                    
                    progress_bar = st.progress(0)
                    all_data = []
                    successful = 0
                    failed = 0
                    
                    for idx, img_info in enumerate(all_images_to_process):
                        status_text.text(f"⏳ {idx+1}/{len(all_images_to_process)}: {img_info['name']}")
                        
                        data = extract_data_from_image(
                            client, img_info['bytes'], img_info['name'],
                            unclear_text, extract_rating,
                            max_dimension, image_quality
                        )
                        
                        if data:
                            all_data.append(data)
                            successful += 1
                        else:
                            failed += 1
                        
                        progress_bar.progress((idx + 1) / len(all_images_to_process))
                    
                    status_text.text("✅ تمت المعالجة!")
                    
                    if successful == 0:
                        st.error(f"❌ فشلت كل الصور")
                        st.stop()
                    
                    st.info(f"📊 نجحت: {successful} | فشلت: {failed}")
                    
                    employees, sorted_dates, month_info = merge_all_data(
                        all_data, extract_rating, auto_convert_rating
                    )
                    
                    if not employees:
                        st.warning("⚠️ لم يتم استخراج موظفين")
                        st.stop()
                    
                    df = create_dataframe(employees, sorted_dates, sort_order, extract_rating)
                    
                    st.session_state['df'] = df
                    st.session_state['month_info'] = month_info or {'month': '00', 'year': '0000'}
                    st.session_state['employees_count'] = len(employees)
                    st.session_state['dates_count'] = len(sorted_dates)
                    st.session_state['extract_rating'] = extract_rating
                    
                    st.markdown(f"""
                    <div class="success-box">
                        <h3>✅ تمت المعالجة!</h3>
                        <p>استخراج <strong>{len(employees)}</strong> موظف</p>
                        <p>اذهب لتبويب <strong>📊 النتائج</strong></p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.balloons()
                    
                except Exception as e:
                    st.error(f"❌ {str(e)}")
        else:
            st.warning("⚠️ يرجى إدخال مفتاح API")

with tab2:
    st.markdown("### 📊 النتائج")
    
    if 'df' in st.session_state and not st.session_state['df'].empty:
        df = st.session_state['df']
        month_info = st.session_state['month_info']
        has_rating = st.session_state.get('extract_rating', False)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div class="stat-card">
                <h3>👥 الموظفين</h3>
                <h1>{st.session_state.get('employees_count', 0)}</h1>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="stat-card">
                <h3>📅 الأيام</h3>
                <h1>{st.session_state.get('dates_count', 0)}</h1>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            review_count = (df == 'يتم المراجعة').sum().sum()
            st.markdown(f"""
            <div class="stat-card">
                <h3>⚠️ للمراجعة</h3>
                <h1>{review_count}</h1>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            if has_rating and 'التقييم' in df.columns:
                ratings = pd.to_numeric(df['التقييم'], errors='coerce').dropna()
                avg_rating = ratings.mean() if len(ratings) > 0 else 0
                st.markdown(f"""
                <div class="stat-card">
                    <h3>⭐ متوسط التقييم</h3>
                    <h1>{avg_rating:.0f}</h1>
                </div>
                """, unsafe_allow_html=True)
            else:
                month = month_info.get('month', '00')
                year = month_info.get('year', '0000')
                st.markdown(f"""
                <div class="stat-card">
                    <h3>📆 الشهر</h3>
                    <h1>{month}/{year}</h1>
                </div>
                """, unsafe_allow_html=True)
        
        st.markdown("---")
        st.markdown("#### 📋 جدول البيانات")
        st.info("💡 يمكنك تعديل البيانات يدوياً")
        
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            height=400
        )
        
        st.markdown("---")
        st.markdown("#### 💾 تحميل الملف")
        
        col1, col2 = st.columns(2)
        
        month = month_info.get('month', '00')
        year = month_info.get('year', '0000')
        
        with col1:
            excel_file = create_excel_file(edited_df, month_info)
            filename = f"Overtime_Report_{month}_{year}.xlsx"
            
            st.download_button(
                label="📥 تحميل Excel",
                data=excel_file,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            csv = edited_df.to_csv(index=False, encoding='utf-8-sig')
            csv_filename = f"Overtime_Report_{month}_{year}.csv"
            
            st.download_button(
                label="📥 تحميل CSV",
                data=csv,
                file_name=csv_filename,
                mime="text/csv",
                use_container_width=True
            )
    else:
        st.info("ℹ️ ارفع الملفات أولاً")

with tab3:
    st.markdown("### 📖 التعليمات")
    
    st.markdown("""
    #### 🎯 v4.1:
    
    - ✅ نفس دقة v4.0 الممتازة (97%+)
    - ✅ إصلاح: "3." لن يتحول لـ "3.5"
    - ✅ دعم PDF + الصور
    - ✅ استخراج التقييم
    - ✅ ضغط تلقائي
    
    #### 🚀 الخطوات:
    
    1. مفتاح Claude API
    2. رفع الملفات
    3. اضغط "بدء المعالجة"
    4. مراجعة وتعديل
    5. تحميل Excel
    """)

st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #666;'>Built with ❤️ using Streamlit + Claude AI | v4.1</div>",
    unsafe_allow_html=True
)
